from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def write_info_three_akadem_hour_by_1_2_3(ws, row_job, col_job, Items, count_akadem_hour):
    for rowes in ws.rows:
        for cell in rowes:
            if count_akadem_hour == 2:
                for row in range(row_job, row_job + 6):
                    for col in range(col_job, col_job + 4):
                        char = get_column_letter(col)
                        if ws[char + str(row)].value is None:
                            if isinstance(ws[char+ str(row)], MergedCell):
                                for x in ws.merged_cells.ranges:
                                    if char + str(row) in x:
                                        if len(Items) > 0:
                                            if Items[-1] != x.start_cell.value:
                                                if x.start_cell.value is None:
                                                    continue
                                                else:
                                                    Items.append(x.start_cell.value)
                                        else:
                                            if x.start_cell.value is None:
                                                continue
                                            else:
                                                Items.append(x.start_cell.value)
                        else:
                            if len(Items) > 0:
                                if Items[-1] != ws[char+str(row)].value:
                                    if ws[char+str(row)].value is None:
                                        continue
                                    else:
                                        Items.append(ws[char+str(row)].value)

                            else:
                                if ws[char + str(row)].value is None:
                                    continue
                                else:
                                    Items.append(ws[char+str(row)].value)
                return Items
            elif count_akadem_hour == 3:
                for row in range(row_job, row_job + 9):
                    for col in range(col_job, col_job + 4):
                        char = get_column_letter(col)
                        if ws[char + str(row)].value is None:
                            if isinstance(ws[char + str(row)], MergedCell):
                                for x in ws.merged_cells.ranges:
                                    if char + str(row) in x:
                                        if len(Items) > 0:
                                            if Items[-1] != x.start_cell.value:
                                                if x.start_cell.value is None:
                                                    continue
                                                else:
                                                    Items.append(x.start_cell.value)
                                        else:
                                            if x.start_cell.value is None:
                                                continue
                                            else:
                                                Items.append(x.start_cell.value)
                        else:
                            if len(Items) > 0:
                                if Items[-1] != ws[char + str(row)].value:
                                    if ws[char + str(row)].value is None:
                                        continue
                                    else:
                                        Items.append(ws[char + str(row)].value)

                            else:
                                if ws[char + str(row)].value is None:
                                    continue
                                else:
                                    Items.append(ws[char + str(row)].value)
                return Items
            elif count_akadem_hour == 1:
                for row in range(row_job, row_job + 3):
                    for col in range(col_job, col_job + 4):
                        char = get_column_letter(col)
                        if ws[char + str(row)].value is None:
                            if isinstance(ws[char + str(row)], MergedCell):
                                for x in ws.merged_cells.ranges:
                                    if char + str(row) in x:
                                        if len(Items) > 0:
                                            if Items[-1] != x.start_cell.value:
                                                if x.start_cell.value is None:
                                                    continue
                                                else:
                                                    Items.append(x.start_cell.value)
                                        else:
                                            if x.start_cell.value is None:
                                                continue
                                            else:
                                                Items.append(x.start_cell.value)
                        else:
                            if len(Items) > 0:
                                if Items[-1] != ws[char + str(row)].value:
                                    if ws[char + str(row)].value is None:
                                        continue
                                    else:
                                        Items.append(ws[char + str(row)].value)

                            else:
                                if ws[char + str(row)].value is None:
                                    continue
                                else:
                                    Items.append(ws[char + str(row)].value)
                return Items

