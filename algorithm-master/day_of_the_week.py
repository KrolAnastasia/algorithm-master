from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

def day_in_excel(ws, row_job, day_col):
    for rowes in ws.rows:
        for cell in rowes:
            for row in range(row_job, row_job+1):
                for col in range(day_col, day_col+1):
                    char = get_column_letter(col)
                    if ws[char+str(row)].value:
                        return ws[char+str(row)].value
                    else:
                        if isinstance(ws[char+str(row)], MergedCell):
                            for x in ws.merged_cells.ranges:
                                if char + str(row) in x:
                                    day_of_week = str(x.start_cell.value)
                                    return day_of_week