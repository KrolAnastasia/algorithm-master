from openpyxl.utils import get_column_letter
from datetime import *
from math import ceil

def month_day(ws, parity, first_cell, last_cell):
    month = []
    for rowes in ws.rows:
        for cell in rowes:
            const_range_for_day = 18
            # верхняя неделя
            if parity == 1:
                parit = 100
                # первый семестр
                if (cell.value == "Сентябрь" or cell.value == "сентябрь" or cell.value == "СЕНТЯБРЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell + 1), cell.row+(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,9,int(ws[char + str(row)].value))))
                elif (cell.value == "Октябрь" or cell.value == "октябрь" or cell.value == "ОКТЯБРЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell + 1), cell.row+(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,10,int(ws[char + str(row)].value))))
                elif (cell.value == "Ноябрь" or cell.value == "ноябрь" or cell.value == "НОЯБРЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell + 1), cell.row+(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,11,int(ws[char + str(row)].value))))
                elif (cell.value == "Декабрь" or cell.value == "декабрь" or cell.value == "ДЕКАБРЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell + 1), cell.row+(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,12,int(ws[char + str(row)].value))))
                # Второй семестр
                elif (cell.value == "Февраль" or cell.value == "февраль" or cell.value == "ФЕВРАЛЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell + 1), cell.row+(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,2,int(ws[char + str(row)].value))))
                elif (cell.value == "Март" or cell.value == "март" or cell.value == "МАРТ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell + 1), cell.row+(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,3,int(ws[char + str(row)].value))))
                elif (cell.value == "Апрель" or cell.value == "апрель" or cell.value == "АПРЕЛЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell + 1), cell.row+(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,4,int(ws[char + str(row)].value))))
                elif (cell.value == "Май" or cell.value == "май" or cell.value == "МАЙ") and cell.row < parit:
                   for row in range(cell.row+(const_range_for_day*first_cell + 1), cell.row+(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,5,int(ws[char + str(row)].value))))
            # нижняя неделя
            elif parity == 2:
                parit = 100  # ограничение области по поиску
                # первый семестр
                if (cell.value == "Сентябрь" or cell.value == "сентябрь" or cell.value == "СЕНТЯБРЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell)+2, cell.row+ceil(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column + 1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,9,int(ws[char + str(row)].value))))
                elif (cell.value == "Октябрь" or cell.value == "октябрь" or cell.value == "ОКТЯБРЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell)+2, cell.row+ceil(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column + 1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,10,int(ws[char + str(row)].value))))
                elif (cell.value == "Ноябрь" or cell.value == "ноябрь" or cell.value == "НОЯБРЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell)+2, cell.row+ceil(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column + 1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,11,int(ws[char + str(row)].value))))
                elif (cell.value == "Декабрь" or cell.value == "декабрь" or cell.value == "ДЕКАБРЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell)+2, cell.row+ceil(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column + 1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,12,int(ws[char + str(row)].value))))
                # Второй семестр
                elif (cell.value == "Февраль" or cell.value == "февраль" or cell.value == "ФЕВРАЛЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell)+2, cell.row+ceil(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column + 1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,2,int(ws[char + str(row)].value))))
                elif (cell.value == "Март" or cell.value == "март" or cell.value == "МАРТ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell)+2, cell.row+ceil(const_range_for_day*last_cell)):
                        for col in range(cell.column, cell.column + 1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,3,int(ws[char + str(row)].value))))
                elif (cell.value == "Апрель" or cell.value == "апрель" or cell.value == "АПРЕЛЬ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell)+2, cell.row+ceil(const_range_for_day*last_cell) ):
                        for col in range(cell.column, cell.column + 1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,4,int(ws[char + str(row)].value))))
                elif (cell.value == "Май" or cell.value == "май" or cell.value == "МАЙ") and cell.row < parit:
                    for row in range(cell.row+(const_range_for_day*first_cell)+2, cell.row+ceil(const_range_for_day*last_cell) ):
                        for col in range(cell.column, cell.column + 1):
                            char = get_column_letter(col)
                            if ws[char + str(row)].value and str(ws[char + str(row)].value).isnumeric():
                                month.append(str(date(date.today().year,5,int(ws[char + str(row)].value))))

    return month
    moth = []