from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def write_name_groups(ws, row_job, col_job, last_row_job):
    name_group = ""
    for rowes in ws.rows:
        for cell in rowes:
            for row in range(row_job, row_job+1):
                for col in range(col_job, last_row_job):
                    char = get_column_letter(col)
                    if ws[char + str(row)].value:
                        name_group += ws[char + str(row)].value
                        name_group = name_group.replace(' ', '')
                    elif ws[char + str(row)].value is None:
                        if isinstance(char + str(row), MergedCell):
                            for x in ws.merged_cells.ranges:
                                if char + str(row) in x:
                                    name_group += ws[char + str(row)].value
                                    name_group = name_group.replace(' ', '')
            return name_group
