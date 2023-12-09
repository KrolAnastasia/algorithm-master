#from chat_bot_helper.Convert_xls_in_xlsx.convertor import Convert
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell
from Serch_path import *
from write_db import *
from range_by_name_groups import *
from chel_col_equally_1_or_2 import *
from day_of_the_week import day_in_excel
from academ_hour import hour
from month import month_day
from minday_row import day_by_table
from name_grops_return import write_name_groups
class Paser_Schedule:
    def parser_schedule_VolgGTU(path, parity):
        wb = load_workbook(path)
        sheet = wb.sheetnames[0]
        ws = wb[sheet]
        days = const_day_of_the_week(ws)
        hours = const_academ_hour(ws)
        name_groups_const_row = const_row_name_groups(ws)
        for rowes in ws.rows:
            for cell in rowes:
                if cell.column == 2:
                    return ("Done")
                else:
                    for row in range(const_range_by_dec_row(ws, parity), const_range_by_saturday(ws, parity), 3):
                        for col in range(const_range_by_dec_col(ws), table_range_by_name_groups(ws)):
                            Items = []
                            char = get_column_letter(col)
                            name_group = []
                            if ws[char + str(row)].value:
                                if isinstance(ws[get_column_letter(col + 1) + str(row)], MergedCell):
                                    for x in ws.merged_cells.ranges:
                                        if char + str(row) in x:
                                            merged = ""
                                            merged += str(x)
                                            index_last_letter_merged_cell = ""
                                            index_first_letter_merged_cell = ""
                                            index_first_number_merged_cell = ""
                                            index_last_number_merged_cell = ""
                                            for last_letter_merged in range(0, len(merged)):
                                                if merged[last_letter_merged] == ":":
                                                    if not merged[last_letter_merged+1].isnumeric() and not merged[last_letter_merged+2].isnumeric():
                                                        index_last_letter_merged_cell += merged[last_letter_merged+1]
                                                        index_last_letter_merged_cell += merged[last_letter_merged+2]
                                                        break
                                                    elif not merged[last_letter_merged+1].isnumeric() and merged[last_letter_merged+2].isnumeric():
                                                        index_last_letter_merged_cell += merged[last_letter_merged+1]
                                                        break
                                            for first_letter_merged in range(0, len(merged)):
                                                if merged[first_letter_merged] == ":":
                                                    if not merged[0].isnumeric() and not merged[1].isnumeric():
                                                        index_first_letter_merged_cell += merged[0]
                                                        index_first_letter_merged_cell += merged[1]
                                                    elif not merged[0].isnumeric() and merged[1].isnumeric():
                                                        index_first_letter_merged_cell += merged[0]
                                            for first_number in range(0, len(merged)):
                                                if merged[first_number] == ":":
                                                    if merged[first_number-1].isnumeric() and merged[first_number-2].isnumeric() and merged[first_number-3].isnumeric():
                                                        index_first_number_merged_cell += merged[first_number-3]
                                                        index_first_number_merged_cell += merged[first_number-2]
                                                        index_first_number_merged_cell += merged[first_number-1]
                                                    elif merged[first_number-1].isnumeric() and merged[first_number-2].isnumeric():
                                                        index_first_number_merged_cell += merged[first_number-2]
                                                        index_first_number_merged_cell += merged[first_number-1]
                                                    elif merged[first_number-1].isnumeric() and not merged[first_number-2].isnumeric():
                                                        index_first_number_merged_cell += merged[first_number-1]

                                            for last_number in range(0, len(merged)):
                                                if merged[last_number] == ":":
                                                    if merged[-1].isnumeric() and merged[-2].isnumeric() and merged[-3].isnumeric():
                                                        index_last_number_merged_cell += merged[-3]
                                                        index_last_number_merged_cell += merged[-2]
                                                        index_last_number_merged_cell += merged[-1]
                                                    elif merged[-1].isnumeric() and merged[-2].isnumeric():
                                                        index_last_number_merged_cell += merged[-2]
                                                        index_last_number_merged_cell += merged[-1]
                                                    elif merged[last_number-1].isnumeric() and not merged[last_number-2].isnumeric():
                                                        index_last_number_merged_cell += merged[last_number-1]

                                            if int(index_first_number_merged_cell) == int(index_last_number_merged_cell):
                                                less_teach_cab_for_db = write_info_three_akadem_hour_by_1_2_3(ws, row, col, Items, 1)
                                                hour_first_for_db = hour(ws, row, hours)
                                                hour_last_for_db = hour(ws, row + 2, hours)
                                                day_for_db = day_in_excel(ws, row, days)
                                                month_for_db = day_by_table(ws, day_in_excel(ws, row, days), parity)
                                                name_group_for_db = write_name_groups(ws, name_groups_const_row,
                                                                                      column_index_from_string(
                                                                                          index_first_letter_merged_cell),
                                                                                      column_index_from_string(
                                                                                          index_last_letter_merged_cell))
                                                Write(less_teach_cab_for_db, hour_first_for_db, hour_last_for_db,
                                                      day_for_db, month_for_db, name_group_for_db, parity)
                                            elif int(index_first_number_merged_cell) - int(index_last_number_merged_cell) == -3:
                                                less_teach_cab_for_db = write_info_three_akadem_hour_by_1_2_3(
                                                    ws, row, col, Items, 2)
                                                month_for_db = day_by_table(ws, day_in_excel(ws, row, days), parity)
                                                hour_first_for_db = hour(ws, row, hours)
                                                hour_last_for_db = hour(ws, row + 5, hours)
                                                day_for_db = day_in_excel(ws, row, days)
                                                name_group_for_db = write_name_groups(ws, name_groups_const_row,
                                                                                      column_index_from_string(
                                                                                          index_first_letter_merged_cell),
                                                                                      column_index_from_string(
                                                                                          index_last_letter_merged_cell))
                                                Write(less_teach_cab_for_db, hour_first_for_db, hour_last_for_db, day_for_db, month_for_db, name_group_for_db, parity)
                                            elif int(index_first_number_merged_cell) - int(index_last_number_merged_cell) == -6:
                                                month_for_db = day_by_table(ws, day_in_excel(ws, row, days), parity)
                                                less_teach_cab_for_db = write_info_three_akadem_hour_by_1_2_3(ws, row, col, Items, 3)
                                                hour_first_for_db = hour(ws, row, hours)
                                                hour_last_for_db = hour(ws, row + 8, hours)
                                                day_for_db = day_in_excel(ws, row, days)
                                                name_group_for_db = write_name_groups(ws,
                                                                                      name_groups_const_row,
                                                                                      column_index_from_string(
                                                                                          index_first_letter_merged_cell),
                                                                                      column_index_from_string(
                                                                                          index_last_letter_merged_cell))
                                                Write(less_teach_cab_for_db, hour_first_for_db, hour_last_for_db,
                                                      day_for_db, month_for_db, name_group_for_db, parity)


const = ("C:/Users/Anastasia/Desktop/algorithm-master/Schedule")
print(Search_path(const))
#Convert.Search_all_derectories(const)
#if Convert.Search_all_derectories(const) == True:
for y in range(0, len(Search_path(const))):
    for i in range(1, 3):
        print(Paser_Schedule.parser_schedule_VolgGTU(Search_path(const)[y], i))
        print(Search_path(const)[y])
        print('Done')

#path = "Schedule/ОНФАТ/ОН_ФАТ_1_курс.xlsx"
#parity = 1
#Paser_Schedule.parser_schedule_VolgGTU(path, parity)
#parity = 2
#Paser_Schedule.parser_schedule_VolgGTU(path, parity)
#print('Done')