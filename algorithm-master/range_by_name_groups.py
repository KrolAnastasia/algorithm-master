from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def table_range_by_name_groups(ws):
    for rowes in ws.rows:
        for cell in rowes:
            parit = 100
            if cell.value == "Декабрь" or cell.value == "декабрь" or cell.value == "ДЕКАБРЬ" and cell.row < parit:
                for row in range(cell.row, cell.row + 1):
                    for col in range(cell.column,
                                     cell.column + 100):
                        # print(col)
                        char = get_column_letter(col)
                        if not isinstance(ws[char + str(row)], MergedCell) and ws[char + str(row)].border.left.style is\
                                None and ws[char + str(row)].value is None:
                            return col
def const_row_name_groups(ws):
    for rowes in ws.rows:
        for cell in rowes:
            parit = 100
            if cell.value == "Декабрь" or cell.value == "декабрь" or cell.value == "ДЕКАБРЬ" and cell.row < parit:
                return cell.row
def const_range_by_dec_row(ws, parity):#начальная
    for rowes in ws.rows:
        for cell in rowes:
            if parity == 1:
                parit = 100
                if cell.value == "четная" or cell.value == "Четная" or cell.value == "ЧЕТНАЯ" and cell.row < parit:
                    return cell.row + 1
            elif parity == 2:
                parit = 100
                if (cell.value == "нечетная" or cell.value == "Нечетная" or cell.value == "НЕЧЕТНАЯ") and cell.row > parit:
                    return cell.row + 1



def const_range_by_saturday(ws, parity):#конечная
    for rowes in ws.rows:
        for cell in rowes:
            rows = 0
            if parity == 1:
                parit = 100
                if cell.value == "СУББОТА" or cell.value == "суббота" or cell.value == "Суббота" and cell.row < parit:
                    for row in range(cell.row, cell.row + 1):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char+str(row)].value:
                                rows = row + 18
                                return rows
            if parity == 2:
                parit = 200
                if (cell.value == "СУББОТА" or cell.value == "суббота" or cell.value == "Суббота") and cell.row > parit:
                    for row in range(cell.row, cell.row + 1):
                        for col in range(cell.column, cell.column+1):
                            char = get_column_letter(col)
                            if ws[char+str(row)].value:
                                rows = row + 18
                                return rows


def const_range_by_dec_col(ws):#Начальная колонка поиска
    for rowes in ws.rows:
        for cell in rowes:
            parit = 100
            if cell.value == "четная" or cell.value == "Четная" or cell.value == "ЧЕТНАЯ" and cell.row < parit:
                return cell.column + 1


def const_day_of_the_week(ws):
    for rowes in ws.rows:
        for cell in rowes:
            parit = 100
            if cell.value == "Декабрь" or cell.value == "декабрь" or cell.value == "ДЕКАБРЬ" and cell.row < parit:
                for row in range(cell.row, cell.row+1):
                    for col in range(cell.column, cell.column+1):
                        if isinstance(ws[get_column_letter(col + 1)+str(row+2)], MergedCell):
                            for x in ws.merged_cells.ranges:
                                if get_column_letter(col + 1) + str(row + 2) in x:
                                    char = str(x.start_cell.value)
                                    if char == "ПОНЕДЕЛЬНИК" or char == "Понедельник" or char == "понедельник" or char == "ПОНЕДЕЛЬНИК ":
                                        return col + 1
                                    elif (char != "ПОНЕДЕЛЬНИК" and char != "Понедельник" and char != "понедельник" and char != "ПОНЕДЕЛЬНИК ") or char is None:
                                        return col - 4

def const_academ_hour(ws):
    for rowes in ws.rows:
        for cell in rowes:
            parit = 100
            if (cell.value == "четная" or cell.value == "ЧЕТНАЯ" or cell.value == "Четная") and cell.row < parit:
                return cell.column+1
