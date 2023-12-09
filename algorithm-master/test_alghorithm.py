from openpyxl import load_workbook
from range_by_name_groups import *
from day_of_the_week import day_in_excel
from name_grops_return import write_name_groups
from month import month_day
from minday_row import day_by_table
from chel_col_equally_1_or_2 import write_info_three_akadem_hour_by_1_2_3
from academ_hour import hour
from write_db import Write
from peewee import *
from peewee_db_test import *

wb = load_workbook("Schedule/ОН_ФЭВТ/ОН_ФЭВТ_1 курс.xlsx")
sheet = wb.sheetnames[0]
ws = wb[sheet]


def test_range_by_name_group():
    total = table_range_by_name_groups(ws)
    assert total == 47

def test_const_col_day_of_the_week():
    total = const_day_of_the_week(ws)
    assert total == 5

def test_const_academ_hour():# узнали где находятся академические часы, а так же убрали два цикла
    total = const_academ_hour(ws)
    assert total == 7

def test_const_begin_col():#с помощью данного юнит теста, проверили где начинается в определенном файле начальная колонка, а так же убрали ненужные циклы, тем самым оптимизируя код
    total = const_range_by_dec_col(ws)
    assert total == 7

def test_const_end_row():#проверяем где конечная строка считывания, сверили с excel файлом, чтобы знать точные значения, где мы хотим остановиться, слегка подправили алгоритм поиска, результат верный
    total1 = const_range_by_saturday(ws, 1)
    total2 = const_range_by_saturday(ws, 2)
    assert total1 == 120, total2 == 229

def test_const_begin_row():# сверяем с excel файлом, ограничиываем условными значениями(12,121), проверяем, результат верный
    total1 = const_range_by_dec_row(ws, 1)
    total2 = const_range_by_dec_row(ws, 2)
    assert total1 == 12, total2 == 121

def test_const_row_name_group():#для учдачного чтения расписания, мы должны знать, к какой группе относится данное расписание, во всех файлах по разному, но в ОН_ФЭВТ_1 имена груп находятся на 11 строке
    total = const_row_name_groups(ws)
    assert total == 11

def test_day_of_the_week():
    day_week_parity = [day_in_excel(ws, 12, const_day_of_the_week(ws)),
                day_in_excel(ws, 37, const_day_of_the_week(ws)),
                day_in_excel(ws, 54, const_day_of_the_week(ws)),
                day_in_excel(ws, 71, const_day_of_the_week(ws)),
                day_in_excel(ws, 91, const_day_of_the_week(ws)),
                day_in_excel(ws, 112, const_day_of_the_week(ws))] # четная

    day_week_not_parity = [day_in_excel(ws, 129, const_day_of_the_week(ws)),
                day_in_excel(ws, 150, const_day_of_the_week(ws)),
                day_in_excel(ws, 168, const_day_of_the_week(ws)),
                day_in_excel(ws, 189, const_day_of_the_week(ws)),
                day_in_excel(ws, 204, const_day_of_the_week(ws)),
                day_in_excel(ws, 221, const_day_of_the_week(ws))]  # нечетная
    day_week_chek = ["ПОНЕДЕЛЬНИК", "ВТОРНИК", "СРЕДА", "ЧЕТВЕРГ", "ПЯТНИЦА", "СУББОТА"]
    for i in range(1, 3):
        if i == 1:
            for y in range(0, len(day_week_parity)):
                assert day_week_parity[i] == day_week_chek[i]
        else:
            for y in range(0, len(day_week_not_parity)):
                assert day_week_not_parity[i] == day_week_chek[i]

def test_name_group():
    total = write_name_groups(ws, const_row_name_groups(ws), 31, 34)#Маш.графика, суббота 1-2 академический час, имя группы ПрИн-167
    assert total == 'ПрИн-167' #немного изменила заполнение группы, убрала пробелы после имен групп, сократила код в функции

def test_month():
    total = month_day(ws, 1, 1, 2)#переведем даты у вторника(четного) в yyyy-mm-dd вид
    assert total == ['2023-09-12',
                     '2023-09-26',
                     '2023-10-10',
                     '2023-10-24',
                     '2023-11-07',
                     '2023-11-21',
                     '2023-12-05',
                     '2023-12-19']

def test_day_by_table():
    total = day_by_table(ws, day_in_excel(ws, 121, const_day_of_the_week(ws)), 2)#физ.подготовка, нечетная, ИВТ-163,1-2 академический час
    assert total == ['2023-09-04', # снова переводим дату в формат yyyy-mm-dd, day_ba_table, возвращает функцию month_day
                     '2023-09-18',
                     '2023-10-02',
                     '2023-10-16',
                     '2023-10-30',
                     '2023-11-13',
                     '2023-11-27',
                     '2023-12-11',
                     '2023-12-25']

def test_write_les_cab_tech():
    Items1 = []
    Items2 = []
    total = [write_info_three_akadem_hour_by_1_2_3(ws, 193, 11, Items1, 1),#193 строка, нечетная неделя, Мат анализ, колонка "K", первый способ чтения
             write_info_three_akadem_hour_by_1_2_3(ws, 102, 19, Items2, 2)] #102 строка, нечетная неделя, информатика, колонка "S", второй способ чтения
    rezult = [['МАТЕМАТИЧЕСКИЙ АНАЛИЗ', 'доц. Бочкин А.М.', 'В-209'],
              ['ИНФОРМАТИКА', 'Сивашова Е.С.,Асеева С.Д.', 'В-1204']]
    for i in range(len(total)):
        assert total[i] == rezult[i]

def test_academ_hour():
    total = hour(ws, 108, const_academ_hour(ws))#мат анализ, 108 строка, ИВТ-161, 5-6 академический час
    assert total == "5-6"

def test_write_all():
    month = '2023-09-02,2023-09-16,2023-09-30,2023-10-14,2023-10-28,2023-11-11,2023-11-25,2023-12-09,2023-12-23'
    rezult = ["ИНФОРМАТИКА", "В-1204", "ИВТ-163", month, 'СУББОТА', "Сивашова Е.С.,Асеева С.Д. ", '1-2', '3-4', 1]
    Items = []
    total = Write(write_info_three_akadem_hour_by_1_2_3(ws, 102, 19, Items, 2), hour(ws, 102, const_academ_hour(ws)), hour(ws, 102+5, const_academ_hour(ws)),
          day_in_excel(ws, 102, const_day_of_the_week(ws)), day_by_table(ws, day_in_excel(ws, 102, const_day_of_the_week(ws)), 1), write_name_groups(ws, const_row_name_groups(ws), 19, 22), 1)
    assert total == rezult

def test_db():
    rezult = ['Предмет: ЭКОНОМ.ПРЕДПР.', 'Преподаватель: Агиевич Т.Г. ', 'Кабинет: А 601']
    with db:
        schedule_day_week_db_lesson = []
        schedule_for_user = Schedule.select(Schedule.lessons, Schedule.teachers, Schedule.cabinets). \
            where(Schedule.name_groups.contains('ЭМР-253'),
                  Schedule.days_of_the_week_all.contains('ЧЕТВЕРГ'),
                  Schedule.parity == 2,
                  Schedule.lessons.contains('ЭКОНОМ.ПРЕДПР'),
                  Schedule.cabinets.contains('А 601'),
                  Schedule.teachers.contains('Агиевич Т.Г.'))
        if schedule_for_user.exists():
            for note_dic in schedule_for_user:
                schedule_day_week_db_lesson.extend(
                    ["Предмет: " + note_dic.lessons, "Преподаватель: " + note_dic.teachers,
                     "Кабинет: " + note_dic.cabinets])
    total = schedule_day_week_db_lesson
    assert total == rezult

#def test_telebot_rezult():
#